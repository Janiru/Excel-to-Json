from openpyxl import Workbook, load_workbook

wb = load_workbook(filename='C:/Users/Janiru Semitha/Downloads/Ananda college prize giving/students.xlsx')
sh = wb.active
print("{")
print('"Prize Giving 2019": [')
for i in range(1,sh.max_row):
    if(sh.cell(row=i+1, column=1).value != None):
        print("{")
        print('"index":', i, ",")
        print('"name": "', sh.cell(row=i+1, column=1).value,'"', ",")
        
        #if the student has more than 1 award
        count = 2
        Svalue = i
        if(sh.cell(row=Svalue+2, column=1).value == None and i != (sh.max_row-2)):
            print('"prizes": [',end="")
            while(sh.cell(row=Svalue+2, column=1).value == None and i != (sh.max_row-2)):
                print('"',sh.cell(row=Svalue+1, column=2).value,'",',end="")
                Svalue+=1
            print('"',sh.cell(row=Svalue+1, column=2).value,'",',end="")
            print(']')
        else:
            print('"prizes": ','["', sh.cell(row=i+1, column=2   ).value,'"]')

            
        if(i==sh.max_row-2):
            print("}")
        else:
            print("},")
print("]")
print("}")